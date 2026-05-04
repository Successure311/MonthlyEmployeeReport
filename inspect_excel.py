import openpyxl
import xlrd

# Read WorkDurationReport with more rows to find actual data
print("=== WorkDurationReport Jan.xls (all rows) ===")
wb2 = xlrd.open_workbook(
    r'E:\Monthly_Employe_Report\WorkDurationReport Jan.xls')
ws2 = wb2.sheet_by_index(0)
print(f"Total rows: {ws2.nrows}, Total cols: {ws2.ncols}")

# Print all non-empty rows
for r in range(ws2.nrows):
    row_data = [str(ws2.cell_value(r, c)).strip() for c in range(ws2.ncols)]
    non_empty = [v for v in row_data if v]
    if non_empty:
        print(f"Row {r}: {non_empty}")

print("\n=== Att.Analysis_Query.xlsx (with data_only) ===")
wb1 = openpyxl.load_workbook(
    r'E:\Monthly_Employe_Report\Att.Analysis_Query.xlsx', read_only=True, data_only=True)
ws1 = wb1.active
rows = list(ws1.iter_rows(max_row=6, values_only=True))
cols1 = rows[0]
print("Columns:", [str(c) for c in cols1])
for r in rows[1:]:
    print([str(v) for v in r])

# Also check unique employees
employees = set()
for row in ws1.iter_rows(min_row=2, max_col=3, values_only=True):
    if row[2]:
        employees.add(str(row[2]))
print(f"\nUnique employees: {sorted(employees)}")

# Check unique statuses
statuses = set()
for row in ws1.iter_rows(min_row=2, max_col=4, values_only=True):
    if row[3]:
        statuses.add(str(row[3]))
print(f"Unique statuses: {sorted(statuses)}")
wb1.close()
